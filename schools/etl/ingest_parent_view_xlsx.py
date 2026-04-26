#!/usr/bin/env python3
"""Ingest Ofsted Parent View data from XLSX files (2021-2026).

CSV format was discontinued after April 2020.
Downloads XLSX files from gov.uk assets for each release date.
Parses school-level data and computes % positive (Strongly Agree + Agree).

Note: Questions changed in November 2025. This handles both old and new formats.
Requires: openpyxl (pip install openpyxl)
"""

import io
import logging
import os
import re
import urllib.request

import psycopg2
from psycopg2.extras import execute_values

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://property:tbFDrBtIM8wQRkPN2luJo21q@localhost:5432/property",
)

# XLSX download URLs from gov.uk (most recent snapshot per academic year)
PARENT_VIEW_XLSX = [
    ("2025-26", "https://assets.publishing.service.gov.uk/media/69ccfc859b2e6e135502d04e/Accessible_Parent_View_Management_Information_as_at_5_January_2026.ods"),
    ("2024-25", "https://assets.publishing.service.gov.uk/media/6837215e4115cfe5bfaa2cb8/Parent_View_Management_Information_as_at_7_April_2025.xlsx"),
    ("2023-24", "https://assets.publishing.service.gov.uk/media/679367af1e5e5352cb24bf02/Parent_View_Management_Information_as_at_2_September_2024.xlsx"),
    ("2022-23", "https://assets.publishing.service.gov.uk/media/6579d6f8254aaa0010050c99/Parent_View_Management_Information_as_at_4_Sep_2023.xlsx"),
    ("2021-22", "https://assets.publishing.service.gov.uk/media/635906e2d3bf7f0bcfa562ea/Parent_View_Management_Information_as_at_5_Sep_2022.xlsx"),
    ("2020-21", "https://assets.publishing.service.gov.uk/media/617bc46ee90e071983346521/Parent_View_Management_Information_as_at_6_September_2021.xlsx"),
]


def _int(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val) if val == val else None  # NaN check
    s = str(val).strip().replace(",", "")
    if s in ("", "z", "x", "c", "k", "u", "ne", "supp", "na", "low", "-", "."):
        return None
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def _pct(val):
    """Parse percentage value — could be 0.46 (decimal) or 46 (integer %) or '46%'."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if val != val:  # NaN
            return None
        # If value is between 0 and 1, it's a decimal percentage
        if 0 <= val <= 1:
            return int(round(val * 100))
        return int(round(val))
    s = str(val).strip().rstrip("%")
    if s in ("", "z", "x", "c", "k", "u", "ne", "supp", "na", "low", "-", "."):
        return None
    try:
        v = float(s)
        if 0 <= v <= 1:
            return int(round(v * 100))
        return int(round(v))
    except (ValueError, TypeError):
        return None


def _add_pcts(a, b):
    if a is None and b is None:
        return None
    return (a or 0) + (b or 0)


def _get_valid_urns(conn):
    with conn.cursor() as cur:
        cur.execute("SELECT urn FROM schools.institutions")
        return {r[0] for r in cur.fetchall()}


def _find_header_row(ws):
    """Find the row containing column headers (look for 'URN')."""
    for row_idx in range(1, min(20, ws.max_row + 1)):
        for col_idx in range(1, min(20, ws.max_column + 1)):
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val and str(cell_val).strip().upper() == "URN":
                return row_idx
    return 1  # fallback to first row


def _col_map(headers):
    """Build a mapping of normalized header → column index."""
    mapping = {}
    for idx, h in enumerate(headers):
        if h:
            mapping[str(h).strip()] = idx
    return mapping


def _find_col(col_map, candidates):
    """Find column index by trying multiple name candidates (case-insensitive substring)."""
    for candidate in candidates:
        for col_name, idx in col_map.items():
            if candidate.lower() in col_name.lower():
                return idx
    return None


def parse_xlsx(data, academic_year, file_ext="xlsx"):
    """Parse XLSX or ODS file data into parent view rows."""
    try:
        if file_ext == "ods":
            # Use odfpy for ODS files
            from odf.opendocument import load as odf_load
            from odf.table import Table, TableRow, TableCell
            from odf.text import P

            doc = odf_load(io.BytesIO(data))
            sheets = doc.getElementsByType(Table)
            if not sheets:
                logger.warning("No sheets found in ODS file")
                return []

            # Find the school-level sheet
            target_sheet = None
            for sheet in sheets:
                name = sheet.getAttribute("name") or ""
                if "school" in name.lower() or "level" in name.lower():
                    target_sheet = sheet
                    break
            if not target_sheet:
                target_sheet = sheets[0]

            # Extract rows from ODS
            all_rows_data = []
            for row in target_sheet.getElementsByType(TableRow):
                cells = row.getElementsByType(TableCell)
                row_data = []
                for cell in cells:
                    repeat = int(cell.getAttribute("numbercolumnsrepeated") or 1)
                    p_elements = cell.getElementsByType(P)
                    if p_elements:
                        text = "".join(
                            "".join(c.data for c in p.childNodes if hasattr(c, "data"))
                            for p in p_elements
                        )
                    else:
                        text = ""
                    for _ in range(min(repeat, 100)):  # cap repeats
                        row_data.append(text.strip() if text else None)
                all_rows_data.append(row_data)

            if not all_rows_data:
                return []

            # Find header row
            header_row_idx = 0
            for idx, row_data in enumerate(all_rows_data[:20]):
                for cell in row_data:
                    if cell and str(cell).strip().upper() == "URN":
                        header_row_idx = idx
                        break
                else:
                    continue
                break

            headers = all_rows_data[header_row_idx]
            data_rows = all_rows_data[header_row_idx + 1:]

        else:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)

            # Find the school-level sheet
            target_sheet = None
            for name in wb.sheetnames:
                if "school" in name.lower() or "level" in name.lower():
                    target_sheet = wb[name]
                    break
            if not target_sheet:
                target_sheet = wb.active

            # Read all rows
            all_rows_data = []
            for row in target_sheet.iter_rows(values_only=True):
                all_rows_data.append(list(row))

            if not all_rows_data:
                return []

            # Find header row
            header_row_idx = 0
            for idx, row_data in enumerate(all_rows_data[:20]):
                for cell in row_data:
                    if cell and str(cell).strip().upper() == "URN":
                        header_row_idx = idx
                        break
                else:
                    continue
                break

            headers = all_rows_data[header_row_idx]
            data_rows = all_rows_data[header_row_idx + 1:]

        col_map = _col_map(headers)
        logger.info("Found %d columns, %d data rows for %s", len(headers), len(data_rows), academic_year)
        logger.info("Sample headers: %s", [h for h in headers[:30] if h])

        # Find columns
        urn_idx = _find_col(col_map, ["URN"])
        submissions_idx = _find_col(col_map, ["Submissions", "Number of submissions", "Responses"])

        if urn_idx is None:
            logger.warning("No URN column found for %s", academic_year)
            return []

        # Question columns — try both old format (Q1. My child...) and new XLSX format (Q1 Strongly Agree)
        def _positive_cols(q_number):
            """Find Strongly Agree and Agree columns for a question number like 'Q1'."""
            sa_idx = _find_col(col_map, [f"{q_number} Strongly Agree", f"{q_number} Strongly agree",
                                          f"{q_number}. ", f"{q_number}."])
            a_idx = None
            # For Agree, we need exact "Q1 Agree" not "Q1 Strongly Agree" or "Q1 Disagree"
            for col_name, idx in col_map.items():
                cn = col_name.strip()
                # XLSX format: exact "Q1 Agree"
                if cn == f"{q_number} Agree":
                    a_idx = idx
                    break
                # Old CSV format: "Q1. My child... Agree" (ends with " Agree" but not "Strongly Agree" or "Disagree")
                if cn.startswith(f"{q_number}.") and cn.endswith(" Agree") and "Strongly" not in cn and "Dis" not in cn:
                    a_idx = idx
                    break
            # For SA, also do exact match
            if sa_idx is None:
                for col_name, idx in col_map.items():
                    cn = col_name.strip()
                    if cn == f"{q_number} Strongly Agree":
                        sa_idx = idx
                        break
            return sa_idx, a_idx

        happy_sa, happy_a = _positive_cols("Q1")
        safe_sa, safe_a = _positive_cols("Q2")
        behaviour_sa, behaviour_a = _positive_cols("Q3")
        bullying_sa, bullying_a = _positive_cols("Q4")
        concerns_sa, concerns_a = _positive_cols("Q6")
        expectations_sa, expectations_a = _positive_cols("Q8")
        doing_well_sa, doing_well_a = _positive_cols("Q9")
        communication_sa, communication_a = _positive_cols("Q10")
        curriculum_sa, curriculum_a = _positive_cols("Q11")
        activities_sa, activities_a = _positive_cols("Q12")
        development_sa, development_a = _positive_cols("Q13")

        # Q14 is Yes/No
        recommend_idx = _find_col(col_map, ["Q14", "recommend this school", "would recommend"])

        logger.info("[%s] Column matching: happy_sa=%s happy_a=%s safe_sa=%s recommend=%s",
                     academic_year, happy_sa, happy_a, safe_sa, recommend_idx)

        rows = []
        for row_data in data_rows:
            if not row_data or len(row_data) <= (urn_idx or 0):
                continue
            urn = _int(row_data[urn_idx] if urn_idx < len(row_data) else None)
            if not urn:
                continue

            submissions = _int(row_data[submissions_idx] if submissions_idx is not None and submissions_idx < len(row_data) else None)

            def _get_positive(sa_idx, a_idx, row):
                sa = _pct(row[sa_idx] if sa_idx is not None and sa_idx < len(row) else None)
                a = _pct(row[a_idx] if a_idx is not None and a_idx < len(row) else None)
                return _add_pcts(sa, a)

            happy = _get_positive(happy_sa, happy_a, row_data)
            safe = _get_positive(safe_sa, safe_a, row_data)
            behaviour = _get_positive(behaviour_sa, behaviour_a, row_data)
            bullying = _get_positive(bullying_sa, bullying_a, row_data)
            concerns = _get_positive(concerns_sa, concerns_a, row_data)
            expectations = _get_positive(expectations_sa, expectations_a, row_data)
            doing_well = _get_positive(doing_well_sa, doing_well_a, row_data)
            communication = _get_positive(communication_sa, communication_a, row_data)
            curriculum = _get_positive(curriculum_sa, curriculum_a, row_data)
            activities = _get_positive(activities_sa, activities_a, row_data)
            development = _get_positive(development_sa, development_a, row_data)

            recommend = _pct(row_data[recommend_idx] if recommend_idx is not None and recommend_idx < len(row_data) else None)

            rows.append((
                urn, academic_year, submissions,
                happy, safe, behaviour, bullying,
                expectations, doing_well, communication,
                curriculum, recommend,
                concerns, development,
                activities,
            ))

        logger.info("Parsed %d Parent View rows for %s", len(rows), academic_year)
        return rows

    except ImportError as e:
        logger.error("Missing library for %s parsing: %s. Install with: pip install openpyxl odfpy", file_ext, e)
        return []
    except Exception as e:
        logger.error("Failed to parse %s file for %s: %s", file_ext, academic_year, e)
        return []


def download_and_parse():
    all_rows = []

    for academic_year, url in PARENT_VIEW_XLSX:
        logger.info("Downloading Parent View %s from %s", academic_year, url)
        req = urllib.request.Request(url, headers={"User-Agent": "PropertyPulse/1.0"})
        try:
            resp = urllib.request.urlopen(req, timeout=120)
            data = resp.read()
            logger.info("Downloaded %.1f MB", len(data) / 1e6)
        except Exception as e:
            logger.warning("Failed to download %s: %s", academic_year, e)
            continue

        file_ext = "ods" if url.endswith(".ods") else "xlsx"
        rows = parse_xlsx(data, academic_year, file_ext=file_ext)
        all_rows.extend(rows)

    logger.info("Total Parent View rows: %d", len(all_rows))
    return all_rows


def load_parent_view(rows, conn):
    valid_urns = _get_valid_urns(conn)
    rows = [r for r in rows if r[0] in valid_urns]
    logger.info("After FK filter: %d rows with valid URNs", len(rows))
    if not rows:
        return
    with conn.cursor() as cur:
        # Don't delete old CSV-sourced rows — just upsert
        execute_values(
            cur,
            """
            INSERT INTO schools.parent_view (
                urn, academic_year, total_responses,
                happy_at_school, feels_safe, good_behaviour, tackled_bullying,
                challenging_work, well_taught, good_communication,
                wide_curriculum, would_recommend,
                well_looked_after, supported_sen,
                good_leadership
            ) VALUES %s
            ON CONFLICT (urn, academic_year) DO UPDATE SET
                total_responses = EXCLUDED.total_responses,
                happy_at_school = EXCLUDED.happy_at_school,
                feels_safe = EXCLUDED.feels_safe,
                good_behaviour = EXCLUDED.good_behaviour,
                tackled_bullying = EXCLUDED.tackled_bullying,
                challenging_work = EXCLUDED.challenging_work,
                well_taught = EXCLUDED.well_taught,
                good_communication = EXCLUDED.good_communication,
                wide_curriculum = EXCLUDED.wide_curriculum,
                would_recommend = EXCLUDED.would_recommend,
                well_looked_after = EXCLUDED.well_looked_after,
                supported_sen = EXCLUDED.supported_sen,
                good_leadership = EXCLUDED.good_leadership
            """,
            rows,
            page_size=5000,
        )
        conn.commit()
        logger.info("Loaded %d Parent View rows", len(rows))


def main():
    conn = psycopg2.connect(DATABASE_URL)
    try:
        rows = download_and_parse()
        if rows:
            load_parent_view(rows, conn)
    finally:
        conn.close()
    logger.info("Parent View XLSX ingestion complete")


if __name__ == "__main__":
    main()
