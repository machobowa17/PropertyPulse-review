#!/usr/bin/env python3
"""Ingest school finance data into schools.finances.

Source: DfE Financial Benchmarking and Insights Tool (FBIT).
- CFR (Consistent Financial Reporting) XLSX for maintained schools
- AAR (Accounts Returns) XLSX for academies

Downloads XLSX files and parses with openpyxl.
"""

import io
import logging
import os
import urllib.request

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://property:tbFDrBtIM8wQRkPN2luJo21q@localhost:5432/property",
)

FBIT_BASE = "https://financial-benchmarking-and-insights-tool.education.gov.uk/files"

# Download 3 years of data
CFR_FILES = {
    "2023-24": f"{FBIT_BASE}/CFR_2023-24_Full_Data_Workbook.xlsx",
    "2022-23": f"{FBIT_BASE}/CFR_2022-23_Full_Data_Workbook.xlsx",
    "2021-22": f"{FBIT_BASE}/CFR_2021-22_Full_Data_Workbook.xlsx",
}

AAR_FILES = {
    "2023-24": f"{FBIT_BASE}/AAR_2023-24_download.xlsx",
    "2022-23": f"{FBIT_BASE}/AAR_2022-23_download.xlsx",
    "2021-22": f"{FBIT_BASE}/AAR_2021-22_download.xlsx",
}


def _to_int(val):
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip().replace(",", "").replace("£", "")
        if val in ("", "z", "x", "c", "k", "DNS", "n/a", "-"):
            return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _to_float(val):
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip().rstrip("%").replace(",", "")
        if val in ("", "z", "x", "c", "k", "DNS", "n/a", "-"):
            return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _get_valid_urns(conn):
    with conn.cursor() as cur:
        cur.execute("SELECT urn FROM schools.institutions")
        return {r[0] for r in cur.fetchall()}


def download_xlsx(url):
    logger.info("Downloading %s", url)
    req = urllib.request.Request(url, headers={"User-Agent": "PropertyPulse/1.0"})
    resp = urllib.request.urlopen(req, timeout=600)
    data = resp.read()
    logger.info("Downloaded %.1f MB", len(data) / 1e6)
    return data


def _find_header_row(sheet, max_rows=10):
    """Find the row containing column headers (look for 'URN' cell)."""
    for r_idx, row in enumerate(sheet.iter_rows(max_row=max_rows, values_only=True)):
        for val in row:
            if val and str(val).strip() == "URN":
                return r_idx
    return None


def _col_index(headers, *names):
    """Find column index by trying multiple header names."""
    for name in names:
        name_lower = name.lower().strip()
        for i, h in enumerate(headers):
            if h and str(h).strip().lower() == name_lower:
                return i
    return None


def parse_cfr(xlsx_data, academic_year, valid_urns):
    """Parse CFR workbook for maintained schools."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_data), read_only=True)

    # Find the data sheet (name varies by year: "CFR Data" vs "CFR_Data")
    sheet = None
    for name in ["CFR Data", "CFR_Data", "Data", "Sheet1"]:
        if name in wb.sheetnames:
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.active

    # Read all rows
    all_rows = list(sheet.iter_rows(values_only=True))

    # Find header row (usually row 0, but older files may have preamble)
    header_idx = 0
    for r_idx, row in enumerate(all_rows[:10]):
        for val in row:
            if val and str(val).strip() == "URN":
                header_idx = r_idx
                break
        if header_idx > 0:
            break

    headers = all_rows[header_idx]
    col_urn = _col_index(headers, "URN")
    col_pupils = _col_index(headers, "No pupils")
    col_income = _col_index(headers, "Total Income: I01:I18 - E30")
    col_expenditure = _col_index(headers, "Total Expenditure: (E01:E29 + E31 + E32)")
    col_staff = _col_index(headers, "Staff Total: (E01:E03) + E05 + (E07: E11) + E26")
    col_balance = _col_index(headers, "In-year Balance: Total Income (I01:I18 - E30) - Total Expenditure (E01:E29 + E31 + E32)")
    col_reserves = _col_index(headers, "Revenue Reserve: B01 + B02 + B06")

    if col_urn is None:
        logger.error("Could not find URN column in CFR sheet. Headers: %s", [h for h in headers if h])
        wb.close()
        return []

    logger.info("CFR column indices: URN=%s, pupils=%s, income=%s, exp=%s, staff=%s, balance=%s, reserves=%s",
                col_urn, col_pupils, col_income, col_expenditure, col_staff, col_balance, col_reserves)

    rows = []
    for row in all_rows[header_idx + 1:]:
        urn = _to_int(row[col_urn] if col_urn is not None and col_urn < len(row) else None)
        if not urn or urn not in valid_urns:
            continue

        pupils = _to_int(row[col_pupils] if col_pupils is not None and col_pupils < len(row) else None)
        total_income = _to_int(row[col_income] if col_income is not None and col_income < len(row) else None)
        total_exp = _to_int(row[col_expenditure] if col_expenditure is not None and col_expenditure < len(row) else None)
        staff_exp = _to_int(row[col_staff] if col_staff is not None and col_staff < len(row) else None)
        balance = _to_int(row[col_balance] if col_balance is not None and col_balance < len(row) else None)
        reserves = _to_int(row[col_reserves] if col_reserves is not None and col_reserves < len(row) else None)

        per_pupil = None
        if total_exp and pupils and pupils > 0:
            per_pupil = int(total_exp / pupils)

        pct_staff = None
        if staff_exp and total_exp and total_exp > 0:
            pct_staff = round(staff_exp / total_exp * 100, 1)

        rows.append((
            urn, academic_year,
            total_income, total_exp, per_pupil,
            staff_exp, balance, reserves, pct_staff,
        ))

    wb.close()
    logger.info("Parsed %d CFR rows for %s", len(rows), academic_year)
    return rows


def parse_aar(xlsx_data, academic_year, valid_urns):
    """Parse AAR workbook for academies."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_data), read_only=True)

    sheet = None
    for name in ["Academies", "Academy", "Data", "Sheet1"]:
        if name in wb.sheetnames:
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.active

    all_rows = list(sheet.iter_rows(values_only=True))

    # AAR has headers in row index 1 (row 0 has codes)
    header_row_idx = None
    for r_idx, row in enumerate(all_rows[:5]):
        for val in row:
            if val and str(val).strip() == "URN":
                header_row_idx = r_idx
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None:
        logger.error("Could not find header row in AAR sheet")
        wb.close()
        return []

    headers = all_rows[header_row_idx]
    col_urn = _col_index(headers, "URN")
    col_pupils = _col_index(headers, "Number of pupils in academy (FTE)")
    col_income = _col_index(headers, "Total Income")
    col_expenditure = _col_index(headers, "Total Expenditure")
    col_staff = _col_index(headers, "Total Staff Costs")
    col_balance = _col_index(headers, "In year balance")
    col_reserves = _col_index(headers, "Revenue Reserve")

    if col_urn is None:
        logger.error("Could not find URN column in AAR sheet")
        wb.close()
        return []

    logger.info("AAR column indices: URN=%s, pupils=%s, income=%s, exp=%s, staff=%s, balance=%s, reserves=%s",
                col_urn, col_pupils, col_income, col_expenditure, col_staff, col_balance, col_reserves)

    rows = []
    for row in all_rows[header_row_idx + 1:]:
        urn = _to_int(row[col_urn] if col_urn is not None and col_urn < len(row) else None)
        if not urn or urn not in valid_urns:
            continue

        pupils = _to_int(row[col_pupils] if col_pupils is not None and col_pupils < len(row) else None)
        total_income = _to_int(row[col_income] if col_income is not None and col_income < len(row) else None)
        total_exp = _to_int(row[col_expenditure] if col_expenditure is not None and col_expenditure < len(row) else None)
        staff_exp = _to_int(row[col_staff] if col_staff is not None and col_staff < len(row) else None)
        balance = _to_int(row[col_balance] if col_balance is not None and col_balance < len(row) else None)
        reserves = _to_int(row[col_reserves] if col_reserves is not None and col_reserves < len(row) else None)

        per_pupil = None
        if total_exp and pupils and pupils > 0:
            per_pupil = int(total_exp / pupils)

        pct_staff = None
        if staff_exp and total_exp and total_exp > 0:
            pct_staff = round(staff_exp / total_exp * 100, 1)

        rows.append((
            urn, academic_year,
            total_income, total_exp, per_pupil,
            staff_exp, balance, reserves, pct_staff,
        ))

    wb.close()
    logger.info("Parsed %d AAR rows for %s", len(rows), academic_year)
    return rows


def load_finances(rows, conn):
    if not rows:
        return
    with conn.cursor() as cur:
        cur.execute("DELETE FROM schools.finances")
        execute_values(
            cur,
            """
            INSERT INTO schools.finances (
                urn, academic_year,
                total_income, total_expenditure, per_pupil_expenditure,
                staff_expenditure, in_year_balance, revenue_reserves,
                pct_budget_staff
            ) VALUES %s
            ON CONFLICT (urn, academic_year) DO UPDATE SET
                total_income = EXCLUDED.total_income,
                total_expenditure = EXCLUDED.total_expenditure,
                per_pupil_expenditure = EXCLUDED.per_pupil_expenditure,
                staff_expenditure = EXCLUDED.staff_expenditure,
                in_year_balance = EXCLUDED.in_year_balance,
                revenue_reserves = EXCLUDED.revenue_reserves,
                pct_budget_staff = EXCLUDED.pct_budget_staff
            """,
            rows,
            page_size=5000,
        )
        conn.commit()
        logger.info("Loaded %d finance rows", len(rows))


def main():
    conn = psycopg2.connect(DATABASE_URL)
    try:
        valid_urns = _get_valid_urns(conn)
        logger.info("Have %d valid URNs", len(valid_urns))

        all_rows = []

        # Process CFR files (maintained schools)
        for year, url in CFR_FILES.items():
            try:
                data = download_xlsx(url)
                rows = parse_cfr(data, year, valid_urns)
                all_rows.extend(rows)
            except Exception as e:
                logger.warning("CFR %s failed: %s", year, e)

        # Process AAR files (academies)
        for year, url in AAR_FILES.items():
            try:
                data = download_xlsx(url)
                rows = parse_aar(data, year, valid_urns)
                all_rows.extend(rows)
            except Exception as e:
                logger.warning("AAR %s failed: %s", year, e)

        if all_rows:
            # Deduplicate: if same URN+year appears in both CFR and AAR, keep one
            seen = set()
            deduped = []
            for r in all_rows:
                key = (r[0], r[1])
                if key not in seen:
                    seen.add(key)
                    deduped.append(r)
            load_finances(deduped, conn)
        else:
            logger.warning("No finance rows parsed from any source")
    finally:
        conn.close()
    logger.info("Finance ingestion complete")


if __name__ == "__main__":
    main()
